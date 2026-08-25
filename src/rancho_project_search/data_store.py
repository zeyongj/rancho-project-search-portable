from __future__ import annotations

import csv
import io
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile
import threading
import unicodedata
from collections.abc import Iterable, Sequence
from dataclasses import dataclass
from datetime import date, datetime
from importlib.resources import files
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

TEXT_FILES = ("pm.csv", "nlm.csv", "fa.csv", "rm.csv", "ap.json", "ar.json")
ALL_FILES = (*TEXT_FILES, "project-list.xlsx")
MAX_UPLOAD_BYTES = 25 * 1024 * 1024

STREET_WORDS = {
    "alley",
    "avenue",
    "ave",
    "boulevard",
    "blvd",
    "circle",
    "court",
    "crescent",
    "cr",
    "drive",
    "dr",
    "highway",
    "hwy",
    "lane",
    "parkway",
    "place",
    "pl",
    "road",
    "rd",
    "street",
    "st",
    "terrace",
    "trail",
    "way",
}


class DataValidationError(ValueError):
    """Raised when an uploaded data file is invalid."""


@dataclass(frozen=True)
class WorkbookRows:
    headers: list[str]
    rows: list[list[Any]]


def normalize_header(value: Any) -> str:
    normalized = unicodedata.normalize("NFKD", str(value or ""))
    return re.sub(r"[^a-z0-9]+", "", normalized.encode("ascii", "ignore").decode().lower())


def project_number(value: Any) -> str:
    match = re.search(r"(?<!\d)(\d{4})(?!\d)", str(value or ""))
    return match.group(1) if match else ""


def display_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value)


def _looks_like_address(line: str) -> bool:
    normalized = re.sub(r"[^a-z0-9]+", " ", line.lower()).strip()
    tokens = normalized.split()
    if not tokens:
        return False
    has_street_word = any(token in STREET_WORDS for token in tokens)
    has_civic_number = bool(re.match(r"^(?:unit\s+)?\d{1,6}(?:\s*[-/]\s*\d{1,6})?\b", normalized))
    has_postal_code = bool(re.search(r"\b[a-z]\d[a-z]\s*\d[a-z]\d\b", line, re.IGNORECASE))
    return has_postal_code or (has_civic_number and has_street_word)


def split_project_name_address(value: Any, city: Any = "") -> tuple[str, str]:
    lines = [re.sub(r"\s+", " ", part).strip(" ,") for part in str(value or "").splitlines()]
    lines = [line for line in lines if line]
    if not lines:
        return "", display_value(city).strip()

    address_start = next((index for index, line in enumerate(lines[1:], start=1) if _looks_like_address(line)), None)
    if address_start is None:
        name_lines = lines[:1]
        address_lines = lines[1:]
    else:
        name_lines = lines[:address_start]
        address_lines = lines[address_start:]

    city_text = re.sub(r"\s+", " ", display_value(city)).strip(" ,")
    address_text = ", ".join(address_lines)
    if city_text and city_text.casefold() not in address_text.casefold():
        address_text = ", ".join(part for part in (address_text, city_text) if part)
    return " – ".join(name_lines), address_text


def _unique_headers(headers: Sequence[Any]) -> list[str]:
    result: list[str] = []
    seen: dict[str, int] = {}
    for index, raw in enumerate(headers, start=1):
        base = display_value(raw).strip() or f"UNNAMED {index}"
        count = seen.get(base, 0) + 1
        seen[base] = count
        result.append(base if count == 1 else f"{base} {count}")
    return result


def _find_column(headers: Sequence[Any], *names: str) -> int | None:
    normalized = [normalize_header(header) for header in headers]
    for name in names:
        target = normalize_header(name)
        if target in normalized:
            return normalized.index(target)
    return None


def _row_value(row: Sequence[Any], column: int | None) -> str:
    if column is None or column >= len(row):
        return ""
    return display_value(row[column]).strip()


class DataStore:
    def __init__(self, data_dir: Path):
        self.data_dir = Path(data_dir).expanduser().resolve()
        self.backup_dir = self.data_dir / "backups"
        self._lock = threading.RLock()

    def initialize(self) -> None:
        self.data_dir.mkdir(parents=True, exist_ok=True)
        defaults = files("rancho_project_search").joinpath("default_data")
        for name in ALL_FILES:
            destination = self.data_dir / name
            if not destination.exists() and defaults.joinpath(name).is_file():
                destination.write_bytes(defaults.joinpath(name).read_bytes())

    def path_for(self, name: str) -> Path:
        if name not in ALL_FILES:
            raise DataValidationError(f"Unsupported data file: {name}")
        return self.data_dir / name

    def list_files(self) -> list[dict[str, Any]]:
        result = []
        for name in ALL_FILES:
            path = self.path_for(name)
            stat = path.stat() if path.exists() else None
            result.append(
                {
                    "name": name,
                    "exists": bool(stat),
                    "size": stat.st_size if stat else 0,
                    "updated": datetime.fromtimestamp(stat.st_mtime).isoformat(timespec="seconds") if stat else None,
                    "editable": name in TEXT_FILES,
                }
            )
        return result

    def read_bytes(self, name: str) -> bytes:
        path = self.path_for(name)
        if not path.exists():
            raise FileNotFoundError(name)
        return path.read_bytes()

    def read_text(self, name: str) -> str:
        if name not in TEXT_FILES:
            raise DataValidationError(f"{name} is not a text data file")
        return self.read_bytes(name).decode("utf-8-sig")

    def replace_text(self, name: str, text: str) -> dict[str, Any]:
        if name not in TEXT_FILES:
            raise DataValidationError(f"{name} cannot be edited as text")
        payload = text.encode("utf-8")
        self._validate_payload(name, payload)
        backups = self._replace_many({name: payload})
        return {"saved": name, "bytes": len(payload), "backups": backups}

    def replace_file(self, name: str, payload: bytes) -> dict[str, Any]:
        if name == "project-list.xlsx":
            return self.import_project_list(payload)
        self._validate_payload(name, payload)
        backups = self._replace_many({name: payload})
        return {"saved": name, "bytes": len(payload), "backups": backups}

    def import_project_list(self, payload: bytes) -> dict[str, Any]:
        if not payload or len(payload) > MAX_UPLOAD_BYTES:
            raise DataValidationError("The workbook is empty or larger than 25 MB")
        try:
            workbook = load_workbook(io.BytesIO(payload), read_only=True, data_only=True)
        except Exception as exc:
            raise DataValidationError(f"Unable to read the Excel workbook: {exc}") from exc

        try:
            active = self._worksheet_rows(workbook, "Active Projects")
            nlm = self._worksheet_rows(workbook, "NLM")
        finally:
            workbook.close()

        pm_payload, active_keys = self._build_pm_csv(active)
        nlm_payload, nlm_keys = self._build_nlm_csv(nlm)
        overlap = sorted(active_keys & nlm_keys)
        replacements = {
            "project-list.xlsx": payload,
            "pm.csv": pm_payload,
            "nlm.csv": nlm_payload,
        }
        backups = self._replace_many(replacements)
        return {
            "saved": list(replacements),
            "activeRows": len(active.rows),
            "activeProjects": len(active_keys),
            "nlmRows": len(nlm.rows),
            "nlmProjects": len(nlm_keys),
            "activeWinsOverNlm": overlap,
            "backups": backups,
        }

    def dataset(self) -> dict[str, Any]:
        with self._lock:
            active = self._project_records("pm.csv", is_nlm=False)
            nlm_source = self._project_records("nlm.csv", is_nlm=True)
            active_keys = {record["proj"] for record in active}
            nlm_keys = {record["proj"] for record in nlm_source}
            overlap = active_keys & nlm_keys
            nlm = [record for record in nlm_source if record["proj"] not in active_keys]
            fa = self._fa_records()
            ap = self._json_list("ap.json")
            ar = self._json_list("ar.json")
            rm = self._rm_records()
        return {
            "active": active,
            "nlm": nlm,
            "fa": fa,
            "ap": ap,
            "ar": ar,
            "rm": rm,
            "summary": {
                "activeProjects": len(active_keys),
                "nlmProjects": len(nlm_keys - active_keys),
                "nlmSourceProjects": len(nlm_keys),
                "activeWinsOverNlm": sorted(overlap),
                "rmProjects": len({record["key"] for record in rm}),
                "rmRecords": len(rm),
            },
        }

    def open_folder(self) -> None:
        self.initialize()
        if sys.platform == "darwin":
            subprocess.Popen(["open", str(self.data_dir)])
        elif os.name == "nt":
            os.startfile(self.data_dir)  # type: ignore[attr-defined]
        else:
            subprocess.Popen(["xdg-open", str(self.data_dir)])

    def _validate_payload(self, name: str, payload: bytes) -> None:
        if name not in ALL_FILES:
            raise DataValidationError(f"Unsupported data file: {name}")
        if not payload:
            raise DataValidationError("The uploaded file is empty")
        if len(payload) > MAX_UPLOAD_BYTES:
            raise DataValidationError("The uploaded file is larger than 25 MB")
        if name.endswith(".json"):
            try:
                parsed = json.loads(payload.decode("utf-8-sig"))
            except (UnicodeDecodeError, json.JSONDecodeError) as exc:
                raise DataValidationError(f"{name} is not valid JSON: {exc}") from exc
            if not isinstance(parsed, list) or not all(isinstance(item, dict) for item in parsed):
                raise DataValidationError(f"{name} must contain a JSON list of objects")
            return
        if name.endswith(".csv"):
            try:
                rows = list(csv.reader(io.StringIO(payload.decode("utf-8-sig"))))
            except (UnicodeDecodeError, csv.Error) as exc:
                raise DataValidationError(f"{name} is not valid UTF-8 CSV: {exc}") from exc
            if not rows or not any(cell.strip() for cell in rows[0]):
                raise DataValidationError(f"{name} has no header row")
            required = {"pm.csv": "proj", "nlm.csv": "proj", "fa.csv": "project", "rm.csv": "file"}.get(name)
            headers = {normalize_header(cell) for row in rows[:10] for cell in row}
            if required and not any(required in header for header in headers):
                raise DataValidationError(f"{name} does not contain the expected {required.upper()} column")

    def _replace_many(self, replacements: dict[str, bytes]) -> list[str]:
        for name, payload in replacements.items():
            self._validate_payload(name, payload)
        timestamp = datetime.now().strftime("%Y%m%d-%H%M%S-%f")
        backup_batch = self.backup_dir / timestamp
        backup_names: list[str] = []
        with self._lock:
            existing = [self.path_for(name) for name in replacements if self.path_for(name).exists()]
            if existing:
                backup_batch.mkdir(parents=True, exist_ok=True)
                for source in existing:
                    shutil.copy2(source, backup_batch / source.name)
                    backup_names.append(f"{timestamp}/{source.name}")

            staged: list[tuple[Path, Path]] = []
            try:
                for name, payload in replacements.items():
                    destination = self.path_for(name)
                    descriptor, temporary_name = tempfile.mkstemp(prefix=f".{name}.", dir=self.data_dir)
                    temporary = Path(temporary_name)
                    with os.fdopen(descriptor, "wb") as handle:
                        handle.write(payload)
                        handle.flush()
                        os.fsync(handle.fileno())
                    staged.append((temporary, destination))
                for temporary, destination in staged:
                    os.replace(temporary, destination)
            finally:
                for temporary, _ in staged:
                    temporary.unlink(missing_ok=True)
        return backup_names

    @staticmethod
    def _worksheet_rows(workbook: Any, requested_name: str) -> WorkbookRows:
        matching = next((name for name in workbook.sheetnames if name.strip().casefold() == requested_name.casefold()), None)
        if not matching:
            raise DataValidationError(f"Workbook is missing the '{requested_name}' worksheet")
        worksheet = workbook[matching]
        values = [list(row) for row in worksheet.iter_rows(values_only=True)]
        header_index = next(
            (
                index
                for index, row in enumerate(values[:20])
                if _find_column(row, "PROJ #", "PROJECT") is not None
                and _find_column(row, "PROJECT NAME") is not None
            ),
            None,
        )
        if header_index is None:
            raise DataValidationError(f"Could not find the project header row in '{requested_name}'")
        raw_headers = values[header_index]
        last_column = max(
            (
                index
                for index in range(len(raw_headers))
                if any(index < len(row) and row[index] not in (None, "") for row in values[header_index:])
            ),
            default=-1,
        )
        headers = _unique_headers(raw_headers[: last_column + 1])
        project_column = _find_column(headers, "PROJ #", "PROJECT")
        rows = [
            list(row[: last_column + 1]) + [None] * max(0, last_column + 1 - len(row))
            for row in values[header_index + 1 :]
            if project_number(_row_value(row, project_column))
        ]
        return WorkbookRows(headers=headers, rows=rows)

    @staticmethod
    def _csv_bytes(headers: Sequence[str], rows: Iterable[Sequence[Any]]) -> bytes:
        output = io.StringIO(newline="")
        writer = csv.writer(output, lineterminator="\n")
        writer.writerow(headers)
        for row in rows:
            writer.writerow([display_value(value) for value in row])
        return output.getvalue().encode("utf-8-sig")

    def _build_pm_csv(self, active: WorkbookRows) -> tuple[bytes, set[str]]:
        project_column = _find_column(active.headers, "PROJ #", "PROJECT")
        name_column = _find_column(active.headers, "PROJECT NAME")
        strata_column = _find_column(active.headers, "STRATA PLAN")
        pm_column = _find_column(active.headers, "PM")
        if project_column is None or name_column is None:
            raise DataValidationError("Active Projects must contain PROJ # and PROJECT NAME columns")
        selected = [project_column, name_column, strata_column, pm_column]
        rows = [[_row_value(row, column) for column in selected] for row in active.rows]
        keys = {project_number(row[0]) for row in rows if project_number(row[0])}
        return self._csv_bytes(["PROJ #", "PROJECT NAME", "STRATA \nPLAN", "PM"], rows), keys

    def _build_nlm_csv(self, nlm: WorkbookRows) -> tuple[bytes, set[str]]:
        project_column = _find_column(nlm.headers, "PROJ #", "PROJECT")
        if project_column is None:
            raise DataValidationError("NLM must contain a PROJ # column")
        keys = {project_number(_row_value(row, project_column)) for row in nlm.rows}
        return self._csv_bytes(nlm.headers, nlm.rows), {key for key in keys if key}

    def _dict_rows(self, name: str) -> list[dict[str, str]]:
        reader = csv.DictReader(io.StringIO(self.read_text(name)))
        return [dict(row) for row in reader]

    def _project_records(self, name: str, *, is_nlm: bool) -> list[dict[str, Any]]:
        records = []
        for row in self._dict_rows(name):
            keys = list(row)
            project_key = next((key for key in keys if normalize_header(key) in {"proj", "project", "projnumber"}), None)
            project_name_key = next((key for key in keys if normalize_header(key) == "projectname"), None)
            city_key = next((key for key in keys if normalize_header(key) == "city"), None)
            strata_key = next((key for key in keys if "strata" in normalize_header(key)), None)
            pm_key = next((key for key in keys if normalize_header(key) == "pm"), None)
            proj = project_number(row.get(project_key or "", ""))
            if not proj:
                continue
            raw_name = row.get(project_name_key or "", "") or ""
            project_name, address = split_project_name_address(raw_name, row.get(city_key or "", ""))
            records.append(
                {
                    "proj": proj,
                    "projectName": project_name,
                    "address": address,
                    "strataPlan": (row.get(strata_key or "", "") or "").strip(),
                    "pm": ((row.get(pm_key or "", "") or "").splitlines() or [""])[0].strip(),
                    "isNlm": is_nlm,
                }
            )
        return records

    def _fa_records(self) -> list[dict[str, str]]:
        rows = self._dict_rows("fa.csv")
        records = []
        for row in rows:
            project_key = next(iter(row), "")
            fa_key = next((key for key in row if normalize_header(key) == "fa"), None)
            proj = project_number(row.get(project_key, ""))
            if proj:
                records.append({"proj": proj, "fa": (row.get(fa_key or "", "") or "").strip()})
        return records

    def _json_list(self, name: str) -> list[dict[str, Any]]:
        parsed = json.loads(self.read_text(name))
        return parsed if isinstance(parsed, list) else []

    def _rm_records(self) -> list[dict[str, str]]:
        rows = list(csv.reader(io.StringIO(self.read_text("rm.csv"))))
        header_index = next((index for index, row in enumerate(rows[:10]) if any("FILE#" in cell.upper() for cell in row)), None)
        if header_index is None:
            return []
        headers = _unique_headers(rows[header_index])
        result = []
        for values in rows[header_index + 1 :]:
            padded = values + [""] * max(0, len(headers) - len(values))
            row = dict(zip(headers, padded, strict=False))
            file_key = next((key for key in row if normalize_header(key) == "file"), None)
            file_number = (row.get(file_key or "", "") or "").strip()
            if not file_number:
                continue
            result.append(
                {
                    "key": file_number[:4],
                    "fileNumber": file_number,
                    "unit": self._value_by_header(row, "unit"),
                    "streetNumber": self._value_by_header(row, "st"),
                    "street": self._value_by_header(row, "street"),
                    "city": self._value_by_header(row, "city"),
                    "pm": self._value_by_header(row, "pm"),
                    "accountant": self._value_by_header(row, "acct"),
                }
            )
        return result

    @staticmethod
    def _value_by_header(row: dict[str, str], target: str) -> str:
        key = next((key for key in row if normalize_header(key) == target), None)
        return (row.get(key or "", "") or "").strip()
